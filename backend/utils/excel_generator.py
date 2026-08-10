import sys
import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Set stdout to UTF-8 to prevent encoding issues
sys.stdout.reconfigure(encoding='utf-8')

def copy_cell_style(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = openpyxl.styles.Font(
            name=src_cell.font.name,
            size=src_cell.font.size,
            bold=src_cell.font.bold,
            italic=src_cell.font.italic,
            color=src_cell.font.color,
            underline=src_cell.font.underline
        ) if src_cell.font else None
        
        dest_cell.fill = openpyxl.styles.PatternFill(
            fill_type=src_cell.fill.fill_type,
            start_color=src_cell.fill.start_color,
            end_color=src_cell.fill.end_color
        ) if src_cell.fill else None
        
        dest_cell.border = openpyxl.styles.Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom
        ) if src_cell.border else None
        
        dest_cell.alignment = openpyxl.styles.Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            wrap_text=src_cell.alignment.wrap_text,
            shrink_to_fit=src_cell.alignment.shrink_to_fit,
            indent=src_cell.alignment.indent
        ) if src_cell.alignment else None
        
        dest_cell.number_format = src_cell.number_format

def adjust_rows(ws, start_row, template_count, target_count):
    """
    Adjust rows from start_row for target_count rows.
    If target_count > template_count, insert rows.
    If target_count < template_count, delete extra rows.
    """
    diff = target_count - template_count
    if diff > 0:
        # Insert rows below the template rows
        ws.insert_rows(start_row + template_count, diff)
        # Copy styles from the last template row to the new rows
        src_row = start_row + template_count - 1
        for r in range(start_row + template_count, start_row + target_count):
            ws.row_dimensions[r].height = ws.row_dimensions[src_row].height
            for c in range(1, ws.max_column + 1):
                copy_cell_style(ws.cell(src_row, c), ws.cell(r, c))
    elif diff < 0:
        # Delete rows
        ws.delete_rows(start_row + target_count, -diff)

def update_signatures(ws, label_row, writer_name, manager_name, director_name):
    """
    Find where signature labels are, insert empty rows, and type the names.
    """
    # Insert 4 rows below the signature labels to type names
    ws.insert_rows(label_row + 1, 4)
    name_row = label_row + 4
    
    # We will look for cells in label_row and write names below them
    for col in range(1, ws.max_column + 1):
        val = ws.cell(label_row, col).value
        if val:
            val_str = str(val).strip().lower()
            if 'người chấm' in val_str:
                ws.cell(name_row, col).value = writer_name
                ws.cell(name_row, col).font = Font(name="Times New Roman", size=11, bold=True)
                ws.cell(name_row, col).alignment = Alignment(horizontal="center")
            elif 'phụ trách' in val_str:
                ws.cell(name_row, col).value = manager_name
                ws.cell(name_row, col).font = Font(name="Times New Roman", size=11, bold=True)
                ws.cell(name_row, col).alignment = Alignment(horizontal="center")
            elif 'thủ trưởng' in val_str:
                ws.cell(name_row, col).value = director_name
                ws.cell(name_row, col).font = Font(name="Times New Roman", size=11, bold=True)
                ws.cell(name_row, col).alignment = Alignment(horizontal="center")

def main():
    if len(sys.argv) < 3:
        print("Error: Missing arguments. Usage: python excel_generator.py <json_path> <output_path>")
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2]

    # Load data
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    month = data['month']
    year = data['year']
    dept_name = data['department_name']
    writer_name = data['writer_name']
    manager_name = data['manager_name']
    director_name = data['director_name']
    
    report_data = data['reportData']
    employees_report = report_data['data']

    template_path = r"g:\QL cham cong\Chấm công tháng 7.2026.xlsx"
    if not os.path.exists(template_path):
        print(f"Error: Template not found at {template_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(template_path)
    
    import calendar
    month_str = f"{month:02d}"
    last_day = calendar.monthrange(year, month)[1]
    date_sig_str = f"Hải Vân, ngày {last_day:02d} tháng {month_str} năm {year}"

    # =========================================================================
    # SHEET 1: Chấm công
    # =========================================================================
    if 'Chấm công' in wb.sheetnames:
        ws = wb['Chấm công']
        
        # 1. Update Title and Header details
        ws.cell(2, 1).value = f"Bộ phận: {dept_name}"
        ws.cell(5, 1).value = f"THÁNG {month_str} NĂM {year}"
        
        # 2. Adjust rows for employees
        # In template, employees are rows 8 to 14 (7 rows)
        start_row = 8
        template_count = 7
        target_count = len(employees_report)
        adjust_rows(ws, start_row, template_count, target_count)
        
        # 3. Write Employee Data
        for i, emp_rep in enumerate(employees_report):
            r = start_row + i
            emp = emp_rep['employee']
            att = emp_rep['attendance']
            sums = emp_rep['summaries']
            
            # STT and Name
            ws.cell(r, 1).value = i + 1
            ws.cell(r, 2).value = emp['full_name']
            
            # Days (Columns C to AG -> 3 to 33)
            for d in range(1, 32):
                col = 2 + d
                date_str = f"{year}-{month_str}-{d:02d}"
                val = att.get(date_str, '')
                ws.cell(r, col).value = val
                
            # Summaries (AH to AM -> 34 to 39)
            ws.cell(r, 34).value = sums['AH']
            ws.cell(r, 35).value = sums['AI']
            ws.cell(r, 36).value = sums['AJ']
            ws.cell(r, 37).value = sums['AK']
            ws.cell(r, 38).value = sums['AL']
            ws.cell(r, 39).value = sums['AM']
            
        # 4. Total Row
        total_row = start_row + target_count
        ws.cell(total_row, 2).value = f"Tổng cộng: {target_count}"
        
        # Write formulas for Total columns (AH to AM)
        col_letters = ['AH', 'AI', 'AJ', 'AK', 'AL', 'AM']
        for col_let in col_letters:
            col_idx = openpyxl.utils.column_index_from_string(col_let)
            ws.cell(total_row, col_idx).value = f"=SUM({col_let}{start_row}:{col_let}{total_row - 1})"
            
        # 5. Date and Signatures
        # We need to find the signature label row. It was originally row 17, but now it's shifted.
        # Let's search for "Người chấm công" or similar in column A or B near the end
        sig_label_row = None
        for r in range(total_row + 1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and 'Người chấm công' in str(val):
                sig_label_row = r
                break
                
        if sig_label_row:
            # Update date string in the row above
            # Search for cell with date text originally at row 16, column AA/AG
            # We can write to the cell on (sig_label_row - 1), column AA (27)
            ws.cell(sig_label_row - 1, 27).value = date_sig_str
            # Add signature names
            update_signatures(ws, sig_label_row, writer_name, manager_name, director_name)

    # =========================================================================
    # SHEET 2: Chấm công trực
    # =========================================================================
    if 'Chấm công trực' in wb.sheetnames:
        ws = wb['Chấm công trực']
        
        # Update details
        ws.cell(2, 1).value = f"Bộ phận: {dept_name}"
        ws.cell(2, 9).value = f"Tháng {month_str} Năm {year}"
        
        # Filter employees with duty
        duty_employees = [e for e in employees_report if e['duty']['has_duty']]
        
        # Template employees rows 7 to 11 (5 rows)
        start_row = 7
        template_count = 5
        target_count = len(duty_employees)
        adjust_rows(ws, start_row, template_count, target_count)
        
        if target_count > 0:
            for i, emp_rep in enumerate(duty_employees):
                r = start_row + i
                emp = emp_rep['employee']
                att = emp_rep['attendance']
                duty = emp_rep['duty']
                
                # STT and Name
                ws.cell(r, 1).value = i + 1
                ws.cell(r, 2).value = emp['full_name']
                
                # Days: write only duty symbol
                for d in range(1, 32):
                    col = 2 + d
                    date_str = f"{year}-{month_str}-{d:02d}"
                    symbol = att.get(date_str, '')
                    # If it is a duty symbol (like T), write it, otherwise blank
                    if symbol == 'T':
                        # Let's check if holiday, write L, else T
                        # We can fetch this from the calculated duty splits in JSON
                        # Wait, we can just write the symbol if it is a duty, or write L if date is in holidays
                        # Let's check if the date is in holidays
                        # We don't have holidays set directly here, but we can check if duty.holiday > 0
                        # For simple alignment, let's just write T or the symbol. If it's a holiday, let's check
                        # If we want to write 'L' for holiday, we check the holiday set in data
                        is_holiday = date_str in data['reportData'].get('holidays', []) or date_str in data.get('holidays', [])
                        # Since we passed holidays from controller, let's make sure it's in the data
                        # Actually we can just write T, or check if holiday and write L. Let's write T since template uses T mostly
                        ws.cell(r, col).value = 'T'
                    elif symbol in ['Td', 'TD']:
                        ws.cell(r, col).value = 'TD'
                    elif symbol in ['cd', 'CD']:
                        ws.cell(r, col).value = 'cd'
                    elif symbol in ['TTc', 'TTC']:
                        ws.cell(r, col).value = 'TTc'
                    else:
                        ws.cell(r, col).value = ''
                        
                # Splits (AH to AK -> 34 to 37)
                ws.cell(r, 34).value = duty['weekday']
                ws.cell(r, 35).value = duty['weekend']
                ws.cell(r, 36).value = duty['holiday']
                # Total formula
                ws.cell(r, 37).value = f"=SUM(AH{r}:AJ{r})"
        else:
            # No duty employees, write a dummy blank row or leave empty
            pass
            
        # Total Row
        total_row = start_row + max(1, target_count)
        ws.cell(total_row, 2).value = f"Tổng cộng: {target_count}"
        col_letters = ['AH', 'AI', 'AJ', 'AK']
        for col_let in col_letters:
            col_idx = openpyxl.utils.column_index_from_string(col_let)
            ws.cell(total_row, col_idx).value = f"=SUM({col_let}{start_row}:{col_let}{total_row - 1})"
            
        # Signatures
        sig_label_row = None
        for r in range(total_row + 1, ws.max_row + 1):
            val = ws.cell(r, 2).value
            if val and 'Người chấm' in str(val):
                sig_label_row = r
                break
        if sig_label_row:
            # Date row is above sig_label_row
            ws.cell(sig_label_row - 1, 29).value = date_sig_str
            update_signatures(ws, sig_label_row, writer_name, manager_name, director_name)

    # =========================================================================
    # SHEET 3: Độc hại theo lương
    # =========================================================================
    if 'Độc hại theo lương' in wb.sheetnames:
        ws = wb['Độc hại theo lương']
        
        # Update details
        ws.cell(2, 1).value = f"Bộ phận: {dept_name}"
        ws.cell(2, 7).value = f"Tháng {month_str} năm {year}"
        
        # Filter employees with toxic salary
        toxic_salary_employees = [e for e in employees_report if e['employee']['has_toxic_salary']]
        
        # Template employee row is just row 7 (1 row)
        start_row = 7
        template_count = 1
        target_count = len(toxic_salary_employees)
        adjust_rows(ws, start_row, template_count, target_count)
        
        for i, emp_rep in enumerate(toxic_salary_employees):
            r = start_row + i
            emp = emp_rep['employee']
            att = emp_rep['attendance']
            toxic = emp_rep['toxic']
            
            ws.cell(r, 1).value = i + 1
            ws.cell(r, 2).value = emp['full_name']
            
            # Days: copy active symbols +, T, Tc
            for d in range(1, 32):
                col = 2 + d
                date_str = f"{year}-{month_str}-{d:02d}"
                symbol = att.get(date_str, '')
                if symbol in ['+', '-', 'T', 'Tc', 'TTc', 'Td', 'cd']:
                    ws.cell(r, col).value = symbol
                else:
                    ws.cell(r, col).value = ''
                    
            # Total Column AH (34)
            ws.cell(r, 34).value = toxic['salary']
            
        # Total Row
        total_row = start_row + max(1, target_count)
        ws.cell(total_row, 2).value = "Tổng cộng:"
        ws.cell(total_row, 34).value = f"=SUM(AH{start_row}:AH{total_row - 1})"
        
        # Signatures
        sig_label_row = None
        for r in range(total_row + 1, ws.max_row + 1):
            val = ws.cell(r, 2).value
            if val and 'Người chấm' in str(val):
                sig_label_row = r
                break
        if sig_label_row:
            # Date row is above sig_label_row. Originally at column Y/AD
            ws.cell(sig_label_row - 1, 25).value = date_sig_str
            update_signatures(ws, sig_label_row, writer_name, manager_name, director_name)

    # =========================================================================
    # SHEET 4: Độc hại hiện vật
    # =========================================================================
    if 'Độc hại hiện vật' in wb.sheetnames:
        ws = wb['Độc hại hiện vật']
        
        # Update details
        ws.cell(2, 1).value = f"Bộ phận: {dept_name}"
        ws.cell(2, 7).value = f"Tháng {month_str} năm {year}"
        
        # Filter employees with toxic in-kind
        toxic_inkind_employees = [e for e in employees_report if e['employee']['has_toxic_in_kind']]
        
        # Template employee row is row 8 (1 row). Note row 7 is "1 Xuất mức 3"
        start_row = 8
        template_count = 1
        target_count = len(toxic_inkind_employees)
        adjust_rows(ws, start_row, template_count, target_count)
        
        for i, emp_rep in enumerate(toxic_inkind_employees):
            r = start_row + i
            emp = emp_rep['employee']
            att = emp_rep['attendance']
            toxic = emp_rep['toxic']
            
            ws.cell(r, 1).value = i + 1
            ws.cell(r, 2).value = emp['full_name']
            
            # Days: copy active symbols +, T, Tc ONLY on weekdays
            # Weekdays map
            for d in range(1, 32):
                col = 2 + d
                date_str = f"{year}-{month_str}-{d:02d}"
                symbol = att.get(date_str, '')
                
                # Check if it's a weekday and not a holiday
                # Since we calculated in-kind in JSON, we only fill it if it matches
                # To be consistent with reportData preview, let's check:
                # Is it a weekday and not holiday?
                # Date object for weekday check
                import datetime
                date_obj = datetime.date(year, month, d)
                is_wkday = date_obj.weekday() < 5 # Mon-Fri
                
                # Check if the date is in holidays
                # If we don't have holiday list, let's try to get it
                # We can see if it was counted in toxic_in_kind.
                # Actually, we can check if it is not weekend, and not in holiday set
                # Let's check data's holidays:
                is_hol = date_str in data['reportData'].get('holidays', []) or date_str in data.get('holidays', [])
                
                if is_wkday and not is_hol and symbol in ['+', '-', 'T', 'Tc', 'TTc', 'Td', 'cd']:
                    ws.cell(r, col).value = symbol
                else:
                    ws.cell(r, col).value = ''
                    
            # Total Column AH (34)
            ws.cell(r, 34).value = toxic['in_kind']
            
        # Total Row
        total_row = start_row + max(1, target_count)
        ws.cell(total_row, 2).value = "Tổng cộng:"
        ws.cell(total_row, 34).value = f"=SUM(AH{start_row}:AH{total_row - 1})"
        
        # Signatures
        sig_label_row = None
        for r in range(total_row + 1, ws.max_row + 1):
            val = ws.cell(r, 2).value
            if val and 'Người chấm' in str(val):
                sig_label_row = r
                break
        if sig_label_row:
            ws.cell(sig_label_row - 1, 25).value = date_sig_str
            update_signatures(ws, sig_label_row, writer_name, manager_name, director_name)

    # Save output
    wb.save(output_path)
    print("Excel generated successfully at", output_path)

if __name__ == '__main__':
    main()
